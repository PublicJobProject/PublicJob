import pandas as pd  # Pandas 라이브러리 import
from selenium import webdriver  # Selenium에서 웹 드라이버 import
from selenium.webdriver import ChromeOptions  # Chrome 옵션 설정을 위한 import
from selenium.webdriver.chrome.service import Service as ChromeService  # Chrome 드라이버 서비스 설정을 위한 import
from webdriver_manager.chrome import ChromeDriverManager  # Chrome 드라이버 관리자 import
from selenium.webdriver.common.by import By  # 웹 요소 검색을 위한 By import
import CreateMonthFile  # 사용자 정의 모듈 import
import ContentParsing  # 사용자 정의 모듈 import
import time  # 시간 지연을 위한 import
from openpyxl import load_workbook, Workbook  # 엑셀 파일 열기 및 새 파일 생성 위한 openpyxl import
from openpyxl.styles import PatternFill, Font, Border, Side  # 엑셀 셀 스타일링을 위한 openpyxl 스타일 import
from preventSleep import prevent_sleep, allow_sleep

options = ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-automation"])  # Selenium 자동화 방지 설정

class Scrap:
    def __init__(self):
        prevent_sleep()
        self.DataList = []  # 데이터 저장을 위한 리스트 초기화

        self.folderDate = CreateMonthFile.createFile()  # 폴더 생성 날짜 지정
        self.file_path = "C:/RPA/지자체 희망일자리/RPA 관리 리스트_한개시트.xlsx"  # 파일 경로 지정
        self.df = self.read_df_file(self.file_path)  # Excel 파일을 DataFrame으로 읽어오는 함수 호출
        self.driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)  # Chrome 웹 드라이버 설정
        self.driver.implicitly_wait(10)  # 웹 요소를 찾기 위한 암묵적 대기 시간 설정
        
        try:
            for item in range(len(self.df)):
                name = self.df.loc[item, '구청명']  # DataFrame에서 '구청명' 가져오기
                url = self.df.loc[item, '게시판 URL']  # DataFrame에서 '게시판 URL' 가져오기
                success = self.df.loc[item, '성공여부']  # DataFrame에서 '성공여부' 가져오기
                if success == "O":
                    continue
                for i in range(3):
                    GroupResult = self.dataCollect(url, item, name)  # 데이터 수집 함수 호출
                    if GroupResult == "성공":
                        self.df.loc[item, '성공여부'] = "O"  # 성공 여부 업데이트
                        break
                    else:
                        print(f"X: {GroupResult}")  # 실패 결과 출력
                        self.df.loc[item, '성공여부'] = f"X: {GroupResult}"  # 실패 여부 업데이트
                # 엑셀 파일에 성공 여부를 저장하고 스타일을 설정
                self.save_success_status(item)
        finally:
            allow_sleep()

    def dataCollect(self, url, item, name):
        검색어입력Xpath = self.df.loc[item, '검색어입력Xpath']  # DataFrame에서 '검색어입력Xpath' 가져오기
        클릭Xpath = self.df.loc[item, '클릭Xpath']  # DataFrame에서 '클릭Xpath' 가져오기
        게시물Xpath = self.df.loc[item, '게시물Xpath']  # DataFrame에서 '게시물Xpath' 가져오기
        검색어 = self.df.loc[item, '검색어']  # DataFrame에서 '검색어' 가져오기
        검색어List = 검색어.split(";")  # ';'을 기준으로 검색어를 리스트로 변환
        
        # DataFrame에서 각 항목의 XPath 가져오기
        게시물_사업명Xpath = self.df.loc[item, '게시물_사업명Xpath']
        게시물_신청기간Xpath = self.df.loc[item, '게시물_신청기간Xpath']
        게시물_근무지Xpath = self.df.loc[item, '게시물_근무지Xpath']
        게시물_임금조건_보수_Xpath = self.df.loc[item, '게시물_임금조건(보수)Xpath']
        게시물_본문Xpath = self.df.loc[item, '게시물_본문Xpath']
        게시물_등록일Xpath = self.df.loc[item, '게시물_등록일Xpath']
        게시물_문의처Xpath = self.df.loc[item, '게시물_문의처Xpath']
        게시물목록Xpath = self.df.loc[item, '게시물목록Xpath']

        self.driver.get(url)  # 주어진 URL로 이동

        for value in 검색어List:
            try:
                getText = ""
                for i in range(3):
                    self.driver.find_element(By.XPATH, 검색어입력Xpath).clear()  # 검색어 입력란 초기화
                    getText = self.driver.find_element(By.XPATH, 검색어입력Xpath).get_attribute("value")
                    if getText == "":
                        break
            except:
                return "검색어입력Xpath를 찾을 수 없습니다"
            
            try:
                for i in range(3):
                    self.driver.find_element(By.XPATH, 검색어입력Xpath).send_keys(value)  # 검색어 입력
                    self.driver.find_element(By.XPATH, 클릭Xpath).click()
                    getText = self.driver.find_element(By.XPATH, 검색어입력Xpath).get_attribute("value")
                    if getText == value:
                        break
            except:
                return "검색어 입력 실패"
            
            for i in range(1, 11):
                TempList = []  # 임시 리스트 초기화
                Modified게시물Xpath = 게시물Xpath.replace(";", str(i))  # 게시물 XPath의 ';'를 숫자로 대체
                
                TempList = []
                Modified게시물Xpath = 게시물Xpath.replace(";", str(i))
                print("="*50)
                print(f"{name} : [{value}] 검색어의 {i}번째 게시물 입니다")
                try: # 게시물 클릭
                    for i in range(3):
                        self.driver.find_element(By.XPATH, Modified게시물Xpath).click()  # 게시물 클릭 시도
                        getText = self.driver.find_element(By.XPATH, 게시물_본문Xpath).text
                        if getText != "":
                            break
                except:
                    return "Modified게시물Xpath를 찾을 수 없습니다."

                try:
                    for i in range(3):
                        본문GetText = self.driver.find_element(By.XPATH, 게시물_본문Xpath).text  # 게시물 본문 텍스트 가져오기
                        if 본문GetText != "":
                            break
                except:
                    return "본문 가져오기 실패" 

                # 수집할 정보 가져오기
                businessName = self.get_element_text_or_none(게시물_사업명Xpath, '사업명', 본문GetText) # 사업명 가져오기
                period = self.get_element_text_or_none(게시물_신청기간Xpath, '신청기간', 본문GetText) # 신청기간 가져오기
                workPlace = self.get_element_text_or_none(게시물_근무지Xpath, '근무지', 본문GetText) # 근무지 가져오기
                salary = self.get_element_text_or_none(게시물_임금조건_보수_Xpath, '임금조건', 본문GetText) # 임금조건 가져오기
                current_url = self.driver.current_url  # 현재 페이지의 URL 가져오기
                contact = self.get_element_text_or_none(게시물_문의처Xpath, '문의처', 본문GetText) # 문의처 가져오기
                registDate = self.get_element_text_or_none(게시물_등록일Xpath, '등록일', 본문GetText) # 등록일 가져오기
                
                # 임시 리스트에 데이터 추가
                TempList.extend([name, businessName, period, workPlace, salary, current_url, registDate, contact])
                self.DataList.append(TempList)  # 결과 리스트에 임시 리스트 추가
                
                # 게시물 목록으로 돌아가기 시도 (예외처리)
                for i in range(3):
                    try:
                        self.driver.find_element(By.XPATH, 게시물목록Xpath).click() # 목록 버튼 클릭
                    except:
                        continue
                    try:
                        GetText = self.driver.find_element(By.XPATH, Modified게시물Xpath).text # 목록 버튼 클릭 성공 여부 확인을 위해 메인화면 게시물 텍스트 수집
                    except:
                        continue
                    if GetText != None:
                        print("목록 버튼 클릭 성공")
                        break   
        
        self.make_df_file(self.DataList)  # DataFrame으로 변환하여 파일로 저장
        return "성공" 

    def get_element_text_or_none(self, xpath, columnName, mainText):
        for i in range(3):
            try:
                text = self.driver.find_element(By.XPATH, xpath).text  # 웹 요소에서 텍스트 가져오기 시도
            except:
                text = ContentParsing.contentParse(mainText, columnName)  # 실패 시 본문에서 데이터 파싱하기
            if text == None:
                time.sleep(0.5)
            else:
                return text  # 가져온 텍스트 또는 파싱한 결과 반환

    def read_df_file(self, file_path): # 주어진 파일 경로에서 엑셀 파일을 읽어와 DataFrame으로 반환합니다.
        df = pd.read_excel(file_path)
        return df

    def make_df_file(self, DataList): # 주어진 데이터 리스트를 DataFrame으로 변환한 후 지정된 경로에 엑셀 파일로 저장합니다.
        tempcolumns = ['구분', '사업명', '신청기간', '근무지', '임금조건(보수)', 'URL', '등록일', '문의전화']  # 열 이름 설정
        tempdf = pd.DataFrame(DataList, columns=tempcolumns)  # 데이터 리스트를 DataFrame으로 변환
        print(tempdf)  # 변환된 DataFrame 출력
        
        # DataFrame을 엑셀 파일로 저장
        file_path = f'C:/RPA/지자체 희망일자리/{self.folderDate}/{self.folderDate}.xlsx'
        tempdf.to_excel(file_path, index=False)

        # 엑셀 파일 열기 및 스타일 적용
        wb = load_workbook(file_path)
        ws = wb.active
        
        fill = PatternFill(start_color="FAC090", end_color="FAC090", fill_type="solid") # 셀 배경색 설정
        font = Font(bold=True) # 글자 볼드체 설정
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')) # 셀 테두리 설정 (좌, 우, 상, 하)

        # 컬럼 헤더 스타일 설정
        for col in range(1, len(tempcolumns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = border
        
        wb.save(file_path) # 엑셀 파일 저장

    def save_success_status(self, item): # openpyxl을 사용하여 기존 엑셀 파일을 열고, 성공 여부를 업데이트 및 스타일을 설정합니다.
        wb = load_workbook(self.file_path) # 기존 엑셀 파일 열기
        ws = wb.active # 활성 시트 가져오기
        row_index = item + 2 # 엑셀 행 인덱스 설정 (1-based index)
        col_index = self.df.columns.get_loc('성공여부') + 1 # 엑셀 열 인덱스 설정 (1-based index)
        cell = ws.cell(row=row_index, column=col_index) # 해당 셀 가져오기
        cell.value = self.df.loc[item, '성공여부'] # 성공 여부 업데이트

        # 스타일 설정
        if row_index == 2:
            fill = PatternFill(start_color="FAC090", end_color="FAC090", fill_type="solid") # 셀 배경색 설정
            font = Font(bold=True) # 글자 볼드체 설정
            border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin')) # 셀 테두리 설정 (좌, 우, 상, 하)
            
            header_cell = ws.cell(row=1, column=col_index)
            header_cell.fill = fill
            header_cell.font = font
            header_cell.border = border

        wb.save(self.file_path) # 엑셀 파일 저장

    def __del__(self): # Scrap 클래스가 삭제될 때 웹 드라이버를 종료합니다.
        self.driver.quit()

if __name__ == "__main__":
    Jobs = Scrap()  # Scrap 클래스의 인스턴스를 생성하여 실행합니다.
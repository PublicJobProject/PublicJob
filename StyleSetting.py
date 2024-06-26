from openpyxl import load_workbook, Workbook  # 엑셀 파일 열기 및 새 파일 생성 위한 openpyxl import
from openpyxl.styles import PatternFill, Font, Border, Side  # 엑셀 셀 스타일링을 위한 openpyxl 스타일 import

def styleSet(file_path, df):
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()
        wb.active.title = 'Sheet1'

    sheet = wb.active

    fill = PatternFill(start_color="FAC090", end_color="FAC090", fill_type="solid")
    font = Font(bold=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for col in range(1, df.shape[1] + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border

    for row in range(2, df.shape[0] + 2):
        for col in range(1, df.shape[1] + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border

    wb.save(file_path)
import os  # 파일 및 폴더 관리를 위한 os 모듈
from datetime import datetime  # 날짜 및 시간 관리를 위한 datetime 모듈
import openpyxl  # 엑셀 파일 생성 및 관리를 위한 openpyxl 모듈
from openpyxl.styles import Font, PatternFill, Border, Side  # 엑셀 스타일 설정을 위한 클래스 import
import shutil  # 파일 복사를 위한 shutil 모듈

def createFile():
    try:
        # 현재 날짜를 YYYY-MM 형식으로 변환
        currentDate = datetime.now().strftime("%Y-%m")

        # 기본 폴더 경로 설정
        basePath = "C:/RPA/지자체 희망일자리"
        folderPath = os.path.join(basePath, currentDate)
        filePath = os.path.join(folderPath, f"{currentDate}.xlsx")

        # RPA 폴더가 존재하지 않으면 생성
        if not os.path.exists("C:/RPA"):
            os.makedirs("C:/RPA")  # RPA 폴더 생성
            print("RPA 폴더가 생성되었습니다.")

        # 지자체 희망일자리 폴더가 존재하지 않으면 생성
        if not os.path.exists(basePath):
            os.makedirs(basePath)  # 지자체 희망일자리 폴더 생성
            print("지자체 희망일자리 폴더가 생성되었습니다.")

        # 지자체 희망일자리 폴더 안쪽에 RPA 관리 리스트_한개시트.xlsx 파일이 없으면 복사
        rpaFile = os.path.join(basePath, "RPA 관리 리스트_한개시트.xlsx")
        if not os.path.exists(rpaFile):
            # 환경 변수에서 문서 디렉토리 가져오기
            userDocuments = os.path.join(os.environ['USERPROFILE'], 'Documents')
            sourceFile = os.path.join(userDocuments, "RPA 관리 리스트 원본", "RPA 관리 리스트_한개시트.xlsx")
            if os.path.exists(sourceFile):
                shutil.copy(sourceFile, rpaFile)  # 파일 복사
                print(f"{rpaFile} 파일이 복사되었습니다.")
            else:
                print(f"원본 파일이 존재하지 않습니다: {sourceFile}")

        # 날짜 폴더가 존재하지 않으면 생성
        if not os.path.exists(folderPath):
            os.makedirs(folderPath)  # 날짜별 폴더 생성
            print(f"{folderPath} 폴더가 생성되었습니다.")

        # 엑셀 파일이 존재하지 않으면 생성
        if not os.path.exists(filePath):
            workbook = openpyxl.Workbook()  # 새로운 엑셀 워크북 생성
            sheet = workbook.active  # 기본 활성화된 시트 선택

            # 컬럼 헤더 설정
            headers = ["구분", "사업명", "신청기간", "근무지", "임금조건(보수)", "URL", "등록일", "문의전화"]
            sheet.append(headers)  # 시트에 헤더 추가

            # 스타일 설정
            fill = PatternFill(start_color="FAC090", end_color="FAC090", fill_type="solid")  # 셀 배경색 설정
            font = Font(bold=True)  # 글자 볼드체 설정
            border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))  # 셀 테두리 설정 (좌, 우, 상, 하)

            # 헤더 셀에 스타일 적용
            for cell in sheet[1]:
                cell.fill = fill  # 배경색 적용
                cell.font = font  # 볼드체 적용
                cell.border = border  # 테두리 적용

            # 엑셀 파일 저장
            workbook.save(filePath)
            print(f"{filePath} 파일이 생성되었습니다.")
        else:
            print(f"{filePath} 파일이 이미 존재합니다.")

    except Exception as e:
        print(f"오류 발생: {str(e)}")

# 폴더 생성 및 파일 생성 예제 실행
createFile()